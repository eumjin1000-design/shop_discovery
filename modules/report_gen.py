"""Step 8 - write the run's findings to an Excel workbook in ./output.

Sheets
------
    Summary    - category, decision, total score, one-line verdict
    Scorecard  - the 100-point breakdown, one row per factor
    Keywords   - generated keywords with rationale and volume estimates
    Details    - raw figures from every analysis module

Interface
---------
    write_report(result: PipelineResult, out_dir: str = "output") -> str
        returns the path to the written .xlsx file
"""
from __future__ import annotations

import os
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .models import PipelineResult
from .timez import stamp as kst_stamp

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_DECISION_FILL = {
    "GO": PatternFill("solid", fgColor="C6EFCE"),
    "WATCH": PatternFill("solid", fgColor="FFEB9C"),
    "NO-GO": PatternFill("solid", fgColor="FFC7CE"),
}


def write_report(result: PipelineResult, out_dir: str = "output") -> str:
    os.makedirs(out_dir, exist_ok=True)
    wb = Workbook()
    _summary_sheet(wb.active, result)
    _scorecard_sheet(wb.create_sheet("Scorecard"), result)
    _keywords_sheet(wb.create_sheet("Keywords"), result)
    _details_sheet(wb.create_sheet("Details"), result)

    slug = re.sub(r"[^a-zA-Z0-9]+", "_", result.request.category).strip("_") or "category"
    path = os.path.join(out_dir, f"shop_discovery_{slug}_{kst_stamp()}.xlsx")
    wb.save(path)
    return path


# --------------------------------------------------------------------------
def _header(ws, row: int, *labels: str) -> None:
    for col, label in enumerate(labels, start=1):
        cell = ws.cell(row=row, column=col, value=label)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT


def _autosize(ws, widths: dict[str, int]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def _summary_sheet(ws, r: PipelineResult) -> None:
    ws.title = "Summary"
    v = r.verdict
    ws["A1"] = "Shop Discovery Report"
    ws["A1"].font = Font(bold=True, size=16)
    rows = [
        ("Category", r.request.category),
        ("Target market", r.request.target_market),
        ("Run at", r.finished_at.strftime("%Y-%m-%d %H:%M:%S")),
        ("Decision", v.decision),
        ("Total score", f"{v.total_score:.1f} / 100"),
        ("Verdict", v.summary),
    ]
    for i, (k, val) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        c = ws.cell(row=i, column=2, value=val)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        if k == "Decision":
            c.font = Font(bold=True)
            c.fill = _DECISION_FILL.get(v.decision, PatternFill())
    _autosize(ws, {"A": 16, "B": 90})


def _scorecard_sheet(ws, r: PipelineResult) -> None:
    _header(ws, 1, "Factor", "Score", "Max", "% of max", "Detail")
    for i, line in enumerate(r.verdict.breakdown, start=2):
        pct = line.score / line.max_score if line.max_score else 0.0
        ws.cell(row=i, column=1, value=line.name)
        ws.cell(row=i, column=2, value=round(line.score, 1))
        ws.cell(row=i, column=3, value=line.max_score)
        ws.cell(row=i, column=4, value=f"{pct*100:.0f}%")
        ws.cell(row=i, column=5, value=line.detail).alignment = Alignment(wrap_text=True)
    total_row = len(r.verdict.breakdown) + 2
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=total_row, column=2, value=round(r.verdict.total_score, 1)).font = Font(bold=True)
    ws.cell(row=total_row, column=3, value=100).font = Font(bold=True)
    _autosize(ws, {"A": 28, "B": 8, "C": 8, "D": 10, "E": 70})


def _keywords_sheet(ws, r: PipelineResult) -> None:
    _header(ws, 1, "Keyword", "Est. monthly volume", "Rationale")
    for i, kw in enumerate(r.keywords, start=2):
        ws.cell(row=i, column=1, value=kw.term)
        ws.cell(row=i, column=2, value=kw.est_monthly_volume or "n/a")
        ws.cell(row=i, column=3, value=kw.rationale).alignment = Alignment(wrap_text=True)
    _autosize(ws, {"A": 32, "B": 18, "C": 60})


def _details_sheet(ws, r: PipelineResult) -> None:
    _header(ws, 1, "Module", "Field", "Value")
    rows: list[tuple[str, str, object]] = []
    t, b, rv, it, mg = r.trend, r.bsr, r.review, r.intent, r.margin
    rows += [
        ("Trend", "growth_ratio (YoY)", t.growth_ratio),
        ("Trend", "stability", t.stability),
        ("Trend", "is_seasonal", t.is_seasonal),
        ("Trend", "notes", t.notes),
        ("Amazon BSR", "best_rank", b.best_rank),
        ("Amazon BSR", "median_rank", b.median_rank),
        ("Amazon BSR", "competing_listings", b.competing_listings),
        ("Amazon BSR", "notes", b.notes),
        ("Reviews", "reviews_analyzed", rv.reviews_analyzed),
        ("Reviews", "avg_rating", rv.avg_rating),
        ("Reviews", "negative_ratio", rv.negative_ratio),
        ("Reviews", "top_complaints", "; ".join(rv.top_complaints)),
        ("Reviews", "notes", rv.notes),
        ("Intent", "commercial_intent", it.commercial_intent),
        ("Intent", "problem_awareness", it.problem_awareness),
        ("Intent", "sample_queries", "; ".join(it.sample_queries)),
        ("Intent", "primary_age", it.primary_age),
        ("Intent", "secondary_age", it.secondary_age),
        ("Intent", "age_rationale", it.age_rationale),
        ("Intent", "notes", it.notes),
        ("Margin", "avg_sourcing_cost", mg.avg_sourcing_cost),
        ("Margin", "avg_retail_price", mg.avg_retail_price),
        ("Margin", "shipping_cost", mg.shipping_cost),
        ("Margin", "platform_fees", mg.platform_fees),
        ("Margin", "ad_cost_estimate", mg.ad_cost_estimate),
        ("Margin", "net_margin", mg.net_margin),
        ("Margin", "net_margin_pct", mg.net_margin_pct),
        ("Margin", "notes", mg.notes),
    ]
    for i, (mod, fld, val) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=mod)
        ws.cell(row=i, column=2, value=fld)
        ws.cell(row=i, column=3, value=val).alignment = Alignment(wrap_text=True)
    _autosize(ws, {"A": 14, "B": 22, "C": 80})
