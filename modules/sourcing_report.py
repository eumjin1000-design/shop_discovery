"""Excel writer for the sourcing list (see modules.sourcing).

Also drops a Spark bulk-input ``.txt`` next to the ``.xlsx`` — one line per
unique ``카테고리|서브카테고리|URL`` — to paste into a scraper's bulk tab.

Interface
---------
    write_sourcing_report(result: SourcingResult, shop_name=None, out_dir="output") -> str
        returns the .xlsx path; the .txt sidecar shares the same stem.
"""
from __future__ import annotations

import os
import re
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .report_gen import _HEADER_FILL, _HEADER_FONT
from .sourcing import SourcingResult

_HEADERS = ["#", "서브카테고리", "브랜드(추정)", "상품명", "변형", "Amazon URL(노드·Prime·리뷰순)",
            "예상가격(USD)", "키워드", "ASIN", "리뷰수", "노드ID"]
_WIDTHS = {"A": 5, "B": 24, "C": 18, "D": 38, "E": 14, "F": 56, "G": 14, "H": 32,
           "I": 14, "J": 10, "K": 14}

# "③ Spark 실수집" sheet (only when result.spark_rows is non-empty)
_SPARK_FILL = PatternFill("solid", fgColor="C0001A")
_SPARK_FONT = Font(bold=True, color="FFFFFF")
_SPARK_HEADERS = ["#", "상품명", "ASIN", "Amazon가격", "Shopify판매가", "Shopify정가",
                  "마진USD", "마진율", "별점", "리뷰수", "판매순위"]
_SPARK_WIDTHS = {"A": 4, "B": 45, "C": 12, "D": 11, "E": 13, "F": 11, "G": 10,
                 "H": 8, "I": 7, "J": 10, "K": 22}
_SPARK_FMT = {4: "$#,##0.00", 5: "$#,##0.00", 6: "$#,##0.00", 7: "$#,##0.00",
              8: "0%", 9: "0.0", 10: "#,##0"}
_SPARK_LAST_COL = "K"  # 11 columns


def _add_spark_sheet(wb, result: SourcingResult) -> None:
    rows = list(getattr(result, "spark_rows", ()) or ())
    if not rows:
        return
    ws = wb.create_sheet("③ Spark 실수집")
    ws.merge_cells(f"A1:{_SPARK_LAST_COL}1")
    title = ws.cell(row=1, column=1, value=f"③ Spark 실수집 — {result.category}  ({len(rows)}개)")
    title.font = Font(bold=True, size=13)
    title.alignment = Alignment(horizontal="center")
    ws.merge_cells(f"A2:{_SPARK_LAST_COL}2")
    note = ws.cell(row=2, column=1, value=(
        "Shopify판매가 = Amazon가격 × (1+마진율)  ·  "
        "Shopify정가 = Shopify판매가 ÷ (1−할인율)  ·  마진USD = Shopify판매가 − Amazon가격"))
    note.font = Font(italic=True, size=9, color="666666")
    note.alignment = Alignment(horizontal="center")
    for col, label in enumerate(_SPARK_HEADERS, start=1):
        c = ws.cell(row=3, column=col, value=label)
        c.fill = _SPARK_FILL
        c.font = _SPARK_FONT
        c.alignment = Alignment(horizontal="center")
    for i, r in enumerate(rows, start=1):
        er = 3 + i
        ws.cell(row=er, column=1, value=i)
        ws.cell(row=er, column=2, value=str(r.get("product_name") or ""))
        ws.cell(row=er, column=3, value=str(r.get("asin") or ""))
        ws.cell(row=er, column=4, value=r.get("price_usd"))
        ws.cell(row=er, column=5, value=r.get("shopify_sell"))
        ws.cell(row=er, column=6, value=r.get("shopify_msrp"))
        ws.cell(row=er, column=7, value=r.get("margin_usd"))
        ws.cell(row=er, column=8, value=r.get("margin_rate"))
        ws.cell(row=er, column=9, value=r.get("rating"))
        ws.cell(row=er, column=10, value=r.get("review_count"))
        ws.cell(row=er, column=11, value=str(r.get("sales_rank") or ""))
        for col, fmt in _SPARK_FMT.items():
            ws.cell(row=er, column=col).number_format = fmt
    for col, width in _SPARK_WIDTHS.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A4"


def write_sourcing_report(result: SourcingResult, shop_name: str | None = None,
                          out_dir: str = "output") -> str:
    os.makedirs(out_dir, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Sourcing"

    row0 = 1
    if shop_name:
        ws.cell(row=1, column=1, value=f"Store: {shop_name}").font = Font(bold=True, size=13)
        ws.cell(row=2, column=1, value=result.summary)
        row0 = 4
    else:
        ws.cell(row=1, column=1, value=result.summary)
        row0 = 3

    for col, label in enumerate(_HEADERS, start=1):
        c = ws.cell(row=row0, column=col, value=label)
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT

    for i, r in enumerate(result.rows, start=1):
        row = row0 + i
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=r.subcategory)
        ws.cell(row=row, column=3, value=r.brand or "")
        ws.cell(row=row, column=4, value=r.product_name)
        ws.cell(row=row, column=5, value=r.variant)
        url_cell = ws.cell(row=row, column=6, value=r.amazon_url)
        url_cell.hyperlink = r.amazon_url
        url_cell.style = "Hyperlink"
        ws.cell(row=row, column=7, value=r.est_price)
        ws.cell(row=row, column=8, value=r.keyword)
        ws.cell(row=row, column=9, value=r.asin or "")
        ws.cell(row=row, column=10, value=r.review_count or "")
        ws.cell(row=row, column=11, value=r.amazon_node_id or "")

    for col, width in _WIDTHS.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = f"A{row0 + 1}"
    ws.cell(row=row0, column=1).alignment = Alignment(horizontal="center")

    _add_spark_sheet(wb, result)

    slug = re.sub(r"[^a-zA-Z0-9]+", "_", result.category).strip("_") or "category"
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = os.path.join(out_dir, f"sourcing_{slug}_{stamp}.xlsx")
    wb.save(path)

    # Spark bulk-input sidecar: unique "카테고리|서브카테고리|URL" lines.
    seen: set[str] = set()
    lines: list[str] = []
    for r in result.rows:
        line = f"{result.category}|{r.subcategory}|{r.amazon_url}"
        if line not in seen:
            seen.add(line)
            lines.append(line)
    Path(path).with_suffix(".txt").write_text("\n".join(lines) + "\n", encoding="utf-8")
    return path
