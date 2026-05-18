"""Batch Excel report: one ranking sheet for many analysed categories.

Interface
---------
    write_batch_report(rows, out_dir="output") -> str
        rows: list of {"name", "total", "decision",
                       "breakdown": [[factor, score, max], ...], "summary"}.
        (Order does not matter — the sheet is written best-first.)
        returns the path to the written .xlsx file
"""
from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from .category_ko import translate as ko_translate
from .ranking_modes import no_ad_score, rank_by_no_ad
from .report_gen import _DECISION_FILL, _HEADER_FILL, _HEADER_FONT  # reuse styles
from .timez import stamp as kst_stamp


def _header(ws, *labels: str) -> None:
    for col, label in enumerate(labels, start=1):
        c = ws.cell(row=1, column=col, value=label)
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT


def write_batch_report(rows: list[dict], out_dir: str = "output") -> str:
    os.makedirs(out_dir, exist_ok=True)
    noad_ranks = rank_by_no_ad(rows)
    rows = sorted(rows, key=lambda r: r.get("total", 0), reverse=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Ranking"

    factor_names = [bn for bn, _s, _m in rows[0]["breakdown"]] if rows else []
    _header(ws, "광고순위", "무광고순위", "카테고리(EN)", "카테고리(한글)",
            "광고 총점", "무광고 총점", "판정", *factor_names, "Verdict")

    for rank, r in enumerate(rows, start=1):
        excel_row = rank + 1
        ws.cell(row=excel_row, column=1, value=rank)
        ws.cell(row=excel_row, column=2, value=noad_ranks.get(r["name"], "-"))
        ws.cell(row=excel_row, column=3, value=r["name"])
        ws.cell(row=excel_row, column=4, value=ko_translate(r["name"]))
        ws.cell(row=excel_row, column=5, value=round(r["total"], 1))
        ws.cell(row=excel_row, column=6, value=no_ad_score(r["breakdown"]))
        cell = ws.cell(row=excel_row, column=7, value=r["decision"])
        cell.font = Font(bold=True)
        if r["decision"] in _DECISION_FILL:
            cell.fill = _DECISION_FILL[r["decision"]]
        for j, (_bn, score, _mx) in enumerate(r["breakdown"]):
            ws.cell(row=excel_row, column=8 + j, value=round(score, 1))
        sc = ws.cell(row=excel_row, column=8 + len(factor_names),
                     value=r.get("summary", ""))
        sc.alignment = Alignment(wrap_text=True)

    ws.column_dimensions["A"].width = 9
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 44
    ws.column_dimensions["D"].width = 26
    ws.column_dimensions["E"].width = 11
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 9
    for i in range(len(factor_names)):
        ws.column_dimensions[chr(ord("H") + i)].width = 13
    ws.column_dimensions[chr(ord("H") + len(factor_names))].width = 80
    ws.freeze_panes = "A2"

    path = os.path.join(out_dir, f"shop_discovery_batch_{kst_stamp()}.xlsx")
    wb.save(path)
    return path
